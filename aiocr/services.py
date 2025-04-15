import logging
from time import sleep

from openpyxl import load_workbook

from aiocr.errors import OCRError, OCRRateLimitError

logger = logging.getLogger('main')

class EMU:
    """English Metric Unit"""

    def __init__(self, value):
        self._value = value

    @property
    def millimeters(self):
        return int(self._value / 36000)

    @property
    def points(self):
        return int(self._value / 12700)


class Workbook:
    IMG_MAX_ROW_OFFSET = 0.9

    def __init__(self):
        self._source_path = None
        self._workbook = None
        self.images = {}
        self.texts = {}

    def _extract_images(self):
        self.images = {}
        for sheet in self._workbook.worksheets:
            for image in sheet._images:
                img_row = self._get_image_row(image, sheet)
                self.images[(sheet.title, img_row)] = image._data()

    def _get_image_row(self, image, worksheet):
        """
        Возвращает номер строки, которой принадлежит картинка.
        Если смещение картинки вниз от верха строки больше (IMG_MAX_ROW_OFFSET * 100) %,
        то относим ее к строке ниже.
        """
        row = image.anchor._from.row + 1
        img_offset = EMU(image.anchor._from.rowOff)
        row_height = worksheet.row_dimensions[row].height
        if row_height is not None:
            if img_offset.points > row_height * self.IMG_MAX_ROW_OFFSET:
                row += 1
        return row

    def _insert_texts(self, workbook):
        current_sheet_title = None
        column = 1
        for path, text in self.texts.items():
            sheet = workbook.get_sheet_by_name(path[0])
            if sheet.title != current_sheet_title:
                current_sheet_title = sheet.title
                column = sheet.max_column + 1
            cell = sheet.cell(row=path[1], column=column)
            cell.value = text

    def ocr(self, client):
        error_count = 0
        logger.info('Распознаем текст картинок...')
        for path, img in self.images.items():
            if path in self.texts:
                continue
            logger.info(f'Страница: {path[0]}, строка: {path[1]}')
            try:
                text = client.ocr(img)
                self.texts[path] = text
                logger.info(f'Текст картинки: {text}')
            except OCRRateLimitError as e:
                logger.error(e)
                error_count += 1
                sleep(10)
            except OCRError as e:
                logger.error(e)
                error_count += 1
        logger.info(f'Обработано картинок: {len(self.texts)}, ошибок: {error_count}')
        if error_count > 0:
            logger.info('Нажмите "OCR", чтобы распознать пропущенные картинки')

    def load(self, path):
        self.images = {}
        self.texts = {}
        self._source_path = path
        logger.info(f'Загружаем файл {path}...')
        self._workbook = load_workbook(path)
        logger.info('Файл успешно загружен')
        self._extract_images()
        logger.info(f'Извлечено картинок: {len(self.images)}')

    def save_as(self, target_path):
        wb = load_workbook(self._source_path)
        self._insert_texts(wb)
        wb.save(target_path)
